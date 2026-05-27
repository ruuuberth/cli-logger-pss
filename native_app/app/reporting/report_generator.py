"""Report Generator base class and interfaces"""
from __future__ import annotations

from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Optional
from dataclasses import dataclass
from datetime import datetime


@dataclass
class ReportConfig:
    """Configuration for report generation"""
    title: str
    output_path: Path
    include_timestamp: bool = True
    format: str = "excel"  # excel, csv, json


class ReportGenerator(ABC):
    """Base class for report generators"""

    def __init__(self, config: ReportConfig):
        self.config = config
        self.data: list[dict[str, Any]] = []

    def add_rows(self, rows: list[dict[str, Any]]) -> None:
        """Add rows to the report"""
        self.data.extend(rows)

    @abstractmethod
    def generate(self) -> Path:
        """Generate and save the report. Return the file path."""
        pass

    def _get_output_filename(self, extension: str) -> Path:
        """Generate output filename with optional timestamp"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S") if self.config.include_timestamp else ""
        base_name = self.config.title.replace(" ", "_").lower()
        
        if timestamp:
            filename = f"{base_name}_{timestamp}.{extension}"
        else:
            filename = f"{base_name}.{extension}"
        
        return self.config.output_path / filename


class ExcelReportGenerator(ReportGenerator):
    """Generate reports in Excel format (.xlsx)"""

    def generate(self) -> Path:
        """Generate Excel report"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl no está instalado. Instala con: pip install openpyxl")

        if not self.data:
            raise ValueError("No hay datos para reportar")

        output_file = self._get_output_filename("xlsx")
        output_file.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Get headers from first row
        if self.data:
            headers = list(self.data[0].keys())
            
            # Write headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Write data rows
            for row_idx, row_data in enumerate(self.data, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, 1):
                max_length = len(str(header))
                for row_data in self.data:
                    max_length = max(max_length, len(str(row_data.get(header, ""))))
                ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 50)

        wb.save(output_file)
        return output_file


class CsvReportGenerator(ReportGenerator):
    """Generate reports in CSV format"""

    def generate(self) -> Path:
        """Generate CSV report"""
        import csv

        if not self.data:
            raise ValueError("No hay datos para reportar")

        output_file = self._get_output_filename("csv")
        output_file.parent.mkdir(parents=True, exist_ok=True)

        # Get headers
        headers = list(self.data[0].keys()) if self.data else []

        # Write CSV
        with open(output_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(self.data)

        return output_file


class JsonReportGenerator(ReportGenerator):
    """Generate reports in JSON format"""

    def generate(self) -> Path:
        """Generate JSON report"""
        import json

        if not self.data:
            raise ValueError("No hay datos para reportar")

        output_file = self._get_output_filename("json")
        output_file.parent.mkdir(parents=True, exist_ok=True)

        # Write JSON
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False, default=str)

        return output_file
